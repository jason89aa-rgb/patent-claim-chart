"""Excel Claim Chart Matrix Export."""
import os
from openpyxl import Workbook
from openpyxl.styles import (Font, PatternFill, Alignment, Border, Side,
                              GradientFill)
from openpyxl.utils import get_column_letter

from core.project import ProjectData
from core.region_capture import (capture_for_mappings,
                                 group_mappings_by_region,
                                 build_region_index)
from utils.color_utils import rgb_to_hex, get_text_color
from utils.term_format import split_by_terms, evidence_chunks
from utils.errlog import log_exception

try:
    from openpyxl.cell.rich_text import CellRichText, TextBlock
    from openpyxl.cell.text import InlineFont
    RICH_TEXT = True
except ImportError:      # openpyxl < 3.1
    RICH_TEXT = False

try:
    from openpyxl.drawing.image import Image as XLImage
    from PIL import Image as PILImage
    IMAGE_OK = True
except ImportError:
    IMAGE_OK = False

JUDGMENT_HEX = {
    "일치":    "2E862E",
    "부분일치": "FFA500",
    "불일치":  "CC2222",
    "미판단":  "888888",
}
INTERPRETATION_LABELS = {
    "문언침해": "Lit.",
    "균등론":   "DOE",
    "넓게해석": "Broad",
    "좁게해석": "Narrow",
}
HEADER_FILL = PatternFill("solid", fgColor="2C3E50")
HEADER_FONT = Font(name="맑은 고딕", bold=True, color="FFFFFF", size=10)
BODY_FONT = Font(name="맑은 고딕", size=9)
BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)


def _apply_cell(ws, row, col, value, font=None, fill=None, alignment=None):
    cell = ws.cell(row=row, column=col, value=value)
    if font:
        cell.font = font
    if fill:
        cell.fill = fill
    if alignment:
        cell.alignment = alignment
    cell.border = BORDER
    return cell


def _term_value(text: str, terms: list, chunks=None):
    """매칭 용어를 색상 처리한 셀 값 (openpyxl 3.1+ 리치텍스트).

    chunks가 주어지면 그것을 사용(선행문헌 term_id 기반 색칠)."""
    if chunks is None:
        if not RICH_TEXT or not terms or not text:
            return text
        chunks = split_by_terms(text, terms)
    elif not RICH_TEXT:
        return "".join(c for c, _ in chunks)
    if not any(c for _, c in chunks):
        return "".join(c for c, _ in chunks) if chunks else text
    blocks = []
    for chunk, color in chunks:
        if not chunk:
            continue
        if color:
            font = InlineFont(rFont="맑은 고딕", sz=9, b=True,
                              color=rgb_to_hex(color).lstrip("#"))
            blocks.append(TextBlock(font, chunk))
        else:
            blocks.append(TextBlock(InlineFont(rFont="맑은 고딕", sz=9),
                                    chunk))
    return CellRichText(*blocks)


def _color_maps(data: ProjectData) -> dict:
    colors = {}
    for claim in data.claims:
        for elem in claim.elements:
            colors[elem.element_id] = tuple(elem.color_rgb)
    for t in data.terms:
        colors[t.term_id] = tuple(t.color_rgb)
    return colors


def _insert_image(ws, mappings, colors, anchor: str,
                  max_px: int = 260, terms=None) -> bool:
    """셀 위치에 도면/문장 캡처 이미지 삽입."""
    if not IMAGE_OK:
        return False
    img_path = capture_for_mappings(mappings, colors, terms=terms)
    if not img_path or not os.path.exists(img_path):
        return False
    try:
        with PILImage.open(img_path) as im:
            iw, ih = im.size
        scale = min(max_px / iw, max_px / ih, 1.0)
        img = XLImage(img_path)
        img.width = int(iw * scale)
        img.height = int(ih * scale)
        img.anchor = anchor
        ws.add_image(img)
        return int(ih * scale)
    except Exception as e:
        print(f"[export_excel] image insert error: {e}")
        return False


def _term_legend_sheet(wb, terms: list):
    if not terms:
        return
    ws = wb.create_sheet(title="용어범례")
    _apply_cell(ws, 1, 1, "용어 ID", font=HEADER_FONT, fill=HEADER_FILL,
                alignment=Alignment(horizontal="center"))
    _apply_cell(ws, 1, 2, "용어", font=HEADER_FONT, fill=HEADER_FILL,
                alignment=Alignment(horizontal="center"))
    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["B"].width = 40
    for i, t in enumerate(terms, start=2):
        rgb = tuple(t.color_rgb)
        hexc = rgb_to_hex(rgb).lstrip("#")
        txt = get_text_color(rgb).lstrip("#")
        _apply_cell(ws, i, 1, t.term_id,
                    font=Font(name="맑은 고딕", bold=True, color=txt, size=10),
                    fill=PatternFill("solid", fgColor=hexc),
                    alignment=Alignment(horizontal="center"))
        _apply_cell(ws, i, 2, t.text, font=BODY_FONT)


def export_excel(data: ProjectData, output_path: str) -> bool:
    try:
        wb = Workbook()

        # 표지 시트
        ws_cover = wb.active
        ws_cover.title = "표지"
        ci = data.case_info
        cover_rows = [
            ("출원번호", ci.application_number),
            ("등록번호", ci.registration_number),
            ("우선일", ci.priority_date),
            ("출원일", ci.application_date),
            ("등록일", ci.registration_date),
            ("출원인", ci.applicant),
            ("패밀리", ", ".join(ci.family_patents)),
            ("비고", ci.notes),
        ]
        for r, (label, val) in enumerate(cover_rows, start=2):
            ws_cover.cell(row=r, column=1, value=label).font = Font(
                name="맑은 고딕", bold=True, size=11)
            ws_cover.cell(row=r, column=2, value=val).font = Font(
                name="맑은 고딕", size=11)
        ws_cover.column_dimensions["A"].width = 16
        ws_cover.column_dimensions["B"].width = 40

        terms = data.terms
        colors = _color_maps(data)
        _term_legend_sheet(wb, terms)

        for claim in data.claims:
            ws = wb.create_sheet(title=f"청구항{claim.claim_number}")

            headers = ["구성요소", "청구항 텍스트", "선행문헌", "대응부분",
                       "도면/문장", "판단", "해석강도", "비고"]
            for col_idx, header in enumerate(headers, start=1):
                _apply_cell(ws, 1, col_idx, header,
                            font=HEADER_FONT, fill=HEADER_FILL,
                            alignment=Alignment(horizontal="center",
                                                vertical="center",
                                                wrap_text=True))

            widths = [10, 45, 18, 40, 40, 10, 10, 18]
            for i, w in enumerate(widths, start=1):
                ws.column_dimensions[get_column_letter(i)].width = w

            claim_maps = [m for m in data.mappings
                          if m.claim_number == claim.claim_number]
            by_elem: dict = {}
            for m in claim_maps:
                by_elem.setdefault(m.element_id, []).append(m)
            # 같은 도면 그룹은 이미지를 첫 행에만 넣는다
            region_index = build_region_index(claim_maps)
            rendered_groups: set = set()

            body = Font(name="맑은 고딕", size=9)
            top_wrap = Alignment(wrap_text=True, vertical="top")
            row = 2

            for elem in claim.elements:
                elem_maps = by_elem.get(elem.element_id, [])
                regions = list(group_mappings_by_region(elem_maps).values()) \
                    or [[]]

                elem_rgb = tuple(elem.color_rgb) if elem.color_rgb \
                    else (200, 200, 200)
                bg_hex = rgb_to_hex(elem_rgb).lstrip("#")
                txt_hex = get_text_color(elem_rgb).lstrip("#")
                elem_fill = PatternFill("solid", fgColor=bg_hex)
                elem_font = Font(name="맑은 고딕", bold=True,
                                 color=txt_hex, size=10)

                for idx, region_maps in enumerate(regions):
                    _apply_cell(ws, row, 1,
                                elem.element_id if idx == 0 else "〃",
                                font=elem_font, fill=elem_fill,
                                alignment=Alignment(horizontal="center",
                                                    vertical="top"))
                    # 청구항 텍스트 (매칭 용어 색상)
                    if idx == 0:
                        _apply_cell(ws, row, 2, _term_value(elem.text, terms),
                                    font=body, alignment=top_wrap)
                    else:
                        _apply_cell(ws, row, 2, "", font=body,
                                    alignment=top_wrap)

                    if not region_maps:
                        for c in (3, 4, 5, 7, 8):
                            _apply_cell(ws, row, c, "-" if c == 4 else "",
                                        font=body, alignment=top_wrap)
                        _apply_cell(ws, row, 6, "미판단",
                                    font=Font(name="맑은 고딕", bold=True,
                                              color="888888", size=9),
                                    alignment=Alignment(horizontal="center",
                                                        vertical="top"))
                        ws.row_dimensions[row].height = 30
                        row += 1
                        continue

                    m0 = region_maps[0]
                    gkey, gmembers = region_index.get(
                        m0.mapping_id, (None, region_maps))
                    first_show = gkey not in rendered_groups
                    if gkey is not None:
                        rendered_groups.add(gkey)
                    _apply_cell(ws, row, 3,
                                f"{os.path.basename(m0.doc_path)}\n"
                                f"p.{m0.page + 1}",
                                font=body, alignment=top_wrap)

                    others = sorted({m.element_id for m in gmembers
                                     if m.element_id != elem.element_id})
                    ev = evidence_chunks(m0, terms)
                    if not (m0.extracted_text or "").strip():
                        ev = [("(도면 참조)", None)]
                    if others:
                        ev = ev + [(f"\n[공통 대응: {', '.join(others)}]",
                                    None)]
                    _apply_cell(ws, row, 4,
                                _term_value("", terms, chunks=ev),
                                font=body, alignment=top_wrap)

                    # 도면/문장 이미지 (E열): 같은 도면 그룹은 첫 행에만
                    if first_show:
                        _apply_cell(ws, row, 5, "", font=body,
                                    alignment=top_wrap)
                        img_h = _insert_image(ws, region_maps and gmembers,
                                              colors, f"E{row}", terms=terms)
                    else:
                        _apply_cell(ws, row, 5, "▲ 위 도면과 동일 (함께 표시됨)",
                                    font=body, alignment=top_wrap)
                        img_h = 0

                    j_hex = JUDGMENT_HEX.get(m0.judgment, "888888")
                    _apply_cell(ws, row, 6, m0.judgment,
                                font=Font(name="맑은 고딕", bold=True,
                                          color=j_hex, size=9),
                                alignment=Alignment(horizontal="center",
                                                    vertical="top"))
                    _apply_cell(ws, row, 7,
                                INTERPRETATION_LABELS.get(m0.interpretation,
                                                          m0.interpretation),
                                font=body,
                                alignment=Alignment(horizontal="center",
                                                    vertical="top"))
                    notes = "\n".join(m.note for m in region_maps if m.note)
                    _apply_cell(ws, row, 8, notes, font=body,
                                alignment=top_wrap)

                    # 이미지가 들어간 행은 이미지 높이에 맞춰 확장
                    if img_h:
                        ws.row_dimensions[row].height = max(60, img_h * 0.78)
                    else:
                        ws.row_dimensions[row].height = 60
                    row += 1

            ws.freeze_panes = "A2"

        wb.save(output_path)
        return None
    except Exception as e:
        log_exception("Excel 내보내기")
        return f"{type(e).__name__}: {e}"
